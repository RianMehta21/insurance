#!/usr/bin/env python3
"""
Standalone workbook verifier.

Forces LibreOffice to recalculate every formula in an .xlsx, then scans the
result for error cells. Exits non-zero if any are found, so CI can refuse to
publish a broken workbook.

Why this exists: openpyxl writes formula STRINGS. It never evaluates them. A
workbook can be saved with 7,500 formulas that are all silently broken and
openpyxl will report nothing wrong. The only way to know they work is to make a
real spreadsheet engine calculate them.

Requires LibreOffice:
    Ubuntu/Debian   sudo apt-get install -y libreoffice-calc
    macOS           brew install --cask libreoffice
    Windows         install LibreOffice, then add the program folder to PATH

Usage:
    python verify_workbook.py VHIS_Compare.xlsx
    python verify_workbook.py VHIS_Compare.xlsx --timeout 300 --json
"""

import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import re

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# The seven values Excel and LibreOffice use to signal a broken formula.
ERROR_VALUES = {
    "#REF!":   "deleted or out-of-range reference",
    "#VALUE!": "wrong data type in an argument",
    "#DIV/0!": "division by zero",
    "#N/A":    "lookup found no match",
    "#NAME?":  "unrecognised function or defined name",
    "#NULL!":  "empty intersection of two ranges",
    "#NUM!":   "invalid number for the operation",
}

# StarBasic. calculateAll() forces a full recalculation regardless of the
# document's AutoCalculate setting, which a plain --convert-to does not
# guarantee. This is the whole reason we inject a macro instead of just
# round-tripping the file through soffice.
MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script"
               script:name="Module1" script:language="StarBasic">
Sub RecalculateAndSave
  ThisComponent.calculateAll()
  ThisComponent.store()
  ThisComponent.close(True)
End Sub
</script:module>
"""

def find_soffice():
    for name in ("soffice", "libreoffice"):
        p = shutil.which(name)
        if p:
            return p
    for p in ("/Applications/LibreOffice.app/Contents/MacOS/soffice",
              r"C:\Program Files\LibreOffice\program\soffice.exe"):
        if os.path.exists(p):
            return p
    return None


def install_macro(soffice, profile: Path, timeout=90):
    """Two phases, and the order is not optional.

    LibreOffice must build its OWN profile first (--terminate_after_init).
    A hand-built directory tree looks superficially right but is missing the
    internal registration LibreOffice expects, and the macro then never
    resolves: soffice launches, finds nothing to run, and hangs until killed.
    Only once the real profile exists do we drop the module into it.
    """
    url = profile.as_uri()
    subprocess.run(
        [soffice, "--headless", "--terminate_after_init",
         f"-env:UserInstallation={url}"],
        capture_output=True, timeout=timeout)

    std = profile / "user" / "basic" / "Standard"
    if not std.exists():
        raise RuntimeError(
            "LibreOffice did not create a usable profile. Formulas were NOT "
            "verified. Check that libreoffice-calc is properly installed.")
    (std / "Module1.xba").write_text(MACRO, encoding="utf-8")
    return url


def recalculate(path: Path, timeout: int):
    soffice = find_soffice()
    if not soffice:
        raise RuntimeError(
            "LibreOffice not found on PATH.\n"
            "  Ubuntu/Debian : sudo apt-get install -y libreoffice-calc\n"
            "  macOS         : brew install --cask libreoffice\n"
            "Or run with --no-recalc for a weaker static reference check that "
            "needs no external tools (not recommended for CI).")

    with tempfile.TemporaryDirectory(prefix="verify-lo-") as tmp:
        profile = Path(tmp) / "profile"
        url = install_macro(soffice, profile)

        # Recalculation rewrites in place, so work on a copy and leave the
        # user's file untouched.
        work = Path(tmp) / path.name
        shutil.copy2(path, work)

        cmd = [soffice, "--headless", "--norestore",
               f"-env:UserInstallation={url}",
               "vnd.sun.star.script:Standard.Module1.RecalculateAndSave"
               "?language=Basic&location=application",
               str(work.absolute())]
        if shutil.which("timeout"):
            cmd = ["timeout", str(timeout)] + cmd

        before = work.stat().st_mtime_ns
        proc = subprocess.run(cmd, capture_output=True, timeout=timeout + 15)
        if work.stat().st_mtime_ns == before:
            raise RuntimeError(
                "LibreOffice did not rewrite the file, so nothing was "
                "recalculated. stderr: "
                + proc.stderr.decode(errors="replace")[:300])

        out = path.with_suffix(".recalc.xlsx")
        shutil.copy2(work, out)
        return out


def scan(path: Path, max_locations=100):
    wb = load_workbook(path, data_only=True)
    counts, locations = {}, []
    total_cells = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                total_cells += 1
                if isinstance(v, str) and v in ERROR_VALUES:
                    counts[v] = counts.get(v, 0) + 1
                    if len(locations) < max_locations:
                        locations.append(f"{ws.title}!{cell.coordinate} {v}")
    wb.close()
    return counts, locations, total_cells


def count_formulas(path: Path):
    wb = load_workbook(path, data_only=False)
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    n += 1
    wb.close()
    return n


# Sheet!$A$1:$B$2, Sheet!$A$1, 'Long Name'!$A$1:$A$9
REF_RE = re.compile(
    r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_.]*))!"
    r"\$?([A-Z]{1,3})\$?(\d+)(?::\$?([A-Z]{1,3})\$?(\d+))?")


def static_check(path: Path, max_locations=100):
    """Bounds-check every cross-sheet reference without evaluating anything.

    This is what --no-recalc runs. It cannot prove a formula returns the right
    number - only a real engine can do that - but it does catch the two error
    classes that actually bite when this workbook is regenerated: a reference
    to a sheet that no longer exists (#NAME?/#REF!) and a range that runs past
    the end of the sheet it points at (#REF!), which is exactly what happens
    when a row count changes and a hardcoded range does not follow it.
    """
    wb = load_workbook(path, data_only=False)
    dims = {ws.title: (ws.max_row, ws.max_column) for ws in wb.worksheets}
    problems = []

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                for m in REF_RE.finditer(v):
                    sheet = m.group(1) or m.group(2)
                    if sheet not in dims:
                        problems.append(
                            f"{ws.title}!{cell.coordinate} -> unknown sheet "
                            f"'{sheet}'")
                        continue
                    max_row, max_col = dims[sheet]
                    rows = [int(m.group(4))]
                    cols = [column_index_from_string(m.group(3))]
                    if m.group(6):
                        rows.append(int(m.group(6)))
                        cols.append(column_index_from_string(m.group(5)))
                    if max(rows) > max_row or max(cols) > max_col:
                        problems.append(
                            f"{ws.title}!{cell.coordinate} -> "
                            f"{sheet}!{m.group(0).split('!')[-1]} exceeds "
                            f"{sheet} ({max_row} rows x {max_col} cols)")
                if len(problems) >= max_locations:
                    break
    wb.close()
    return problems


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--timeout", type=int, default=300)
    ap.add_argument("--json", action="store_true")
    ap.add_argument("--keep", action="store_true",
                    help="keep the recalculated copy for inspection")
    ap.add_argument("--no-recalc", action="store_true",
                    help="skip LibreOffice and run the static reference check "
                         "only. Weaker: catches broken references, not wrong "
                         "answers. Not recommended for CI.")
    args = ap.parse_args()

    path = Path(args.workbook)
    if not path.exists():
        print(f"not found: {path}")
        sys.exit(2)

    formulas = count_formulas(path)

    if args.no_recalc:
        problems = static_check(path)
        result = {
            "status": "success" if not problems else "errors_found",
            "mode": "static-only (no recalculation)",
            "workbook": str(path),
            "total_formulas": formulas,
            "total_errors": len(problems),
            "locations": problems,
        }
        if args.json:
            print(json.dumps(result, indent=2))
        else:
            print(f"workbook : {path}")
            print(f"formulas : {formulas:,}")
            print(f"mode     : static reference check, NOT a recalculation")
            print(f"problems : {len(problems)}")
            for p in problems[:20]:
                print("  ", p)
            if not problems:
                print("\nAll cross-sheet references resolve and are in bounds.")
                print("This does NOT prove the formulas compute correct values.")
        sys.exit(0 if not problems else 1)

    recalced = recalculate(path, args.timeout)
    counts, locations, cells = scan(recalced)
    if not args.keep:
        recalced.unlink(missing_ok=True)

    total = sum(counts.values())
    result = {
        "status": "success" if total == 0 else "errors_found",
        "workbook": str(path),
        "total_formulas": formulas,
        "populated_cells": cells,
        "total_errors": total,
        "error_summary": counts,
        "locations": locations,
    }

    if args.json:
        print(json.dumps(result, indent=2))
    else:
        print(f"workbook   : {path}")
        print(f"formulas   : {formulas:,}")
        print(f"cells      : {cells:,}")
        print(f"errors     : {total}")
        if counts:
            print("\nbreakdown:")
            for k, v in sorted(counts.items(), key=lambda x: -x[1]):
                print(f"  {k:<9} {v:>6}   {ERROR_VALUES[k]}")
            print("\nfirst locations:")
            for loc in locations[:20]:
                print("  ", loc)
        else:
            print("\nAll formulas evaluated cleanly.")

    sys.exit(0 if total == 0 else 1)


if __name__ == "__main__":
    main()
