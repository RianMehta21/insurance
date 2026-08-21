#!/usr/bin/env python3
"""
Builds the VHIS comparison workbook.

The workbook is the application, not a report. Change the age on the Client
sheet and all 579 plans reprice instantly via INDEX/MATCH. No Python, no
macros, no add-ins on the user's machine.

Sheets
  Client    inputs (age, gender, smoker, basis) + legend + disclaimers
  Compare   579 plans, AutoFilter on, live premium columns
  Benefits  scraped/override benefit fields, user-editable
  prem      hidden wide premium matrix, one row per cert|basis|gender|smoker
  Notes     provenance, refresh date, assumptions

Usage:
    python make_workbook.py --data-dir current --out VHIS_Compare.xlsx
"""

import argparse
import json
import os
import re
from collections import OrderedDict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# reuse the parsing/loading logic already written and tested
import build_vhis as bv

MAX_AGE = 100
N_AGE_COLS = MAX_AGE + 1                      # ages 0..100
PREM_FIRST_COL = 2                            # column B holds age 0
PREM_LAST_COL = PREM_FIRST_COL + MAX_AGE      # column CX

ARIAL = "Arial"
BLUE = Font(name=ARIAL, size=10, color="0000FF", bold=True)
BLACK = Font(name=ARIAL, size=10)
HDR = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
TITLE = Font(name=ARIAL, size=14, bold=True)
SUB = Font(name=ARIAL, size=10, italic=True, color="595959")
HDR_FILL = PatternFill("solid", fgColor="1F3864")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
EDIT_FILL = PatternFill("solid", fgColor="FFF2CC")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def money(fmt_zero_dash=True):
    return '#,##0;(#,##0);-' if fmt_zero_dash else '#,##0'


# ---------------------------------------------------------------------------
def build_premium_rows(prem_by_cert):
    """One row per cert|basis|gender|smoker. Blank where the plan has no
    premium at that age, so Excel's COUNT can distinguish 'no cover' from 0."""
    rows = []
    for cert, rec in prem_by_cert.items():
        pm = rec.get("premium") or {}
        for b in ("S", "R"):
            for g in ("M", "F"):
                for s in ("N", "Y"):
                    series = pm.get(f"{b}{g}{s}")
                    if not series:
                        continue
                    vals = [series.get(str(a)) for a in range(N_AGE_COLS)]
                    if all(v is None for v in vals):
                        continue
                    rows.append([f"{cert}|{b}|{g}|{s}"] + vals)
    rows.sort(key=lambda r: r[0])
    return rows


def sheet_client(wb, catalog_n, prem_date_max):
    ws = wb.create_sheet("Client", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 82

    ws["B2"] = "VHIS Plan Comparison"
    ws["B2"].font = TITLE
    ws["B3"] = (f"{catalog_n} certified plan variants across 33 insurers. "
                "Change the yellow cells and every plan reprices.")
    ws["B3"].font = SUB

    ws["B5"] = "CLIENT PROFILE"
    ws["B5"].font = Font(name=ARIAL, size=11, bold=True)

    spec = [
        ("Age (years)", 35, "B6",
         "Client's actual age. Each insurer's age basis is applied automatically."),
        ("Gender", "M", "B7", "M or F. 220 of the variants are gender-rated; the rest are unisex."),
        ("Smoker", "N", "B8", "N or Y. Only 46 variants price smokers differently."),
        ("Premium basis", "S", "B9",
         "S = standalone policy. R = rider attached to another policy. Falls back automatically."),
    ]
    r = 6
    for label, default, _, note in spec:
        ws.cell(r, 2, label).font = BLACK
        c = ws.cell(r, 3, default)
        c.font = BLUE
        c.fill = INPUT_FILL
        c.border = BOX
        c.alignment = Alignment(horizontal="center")
        ws.cell(r, 4, note).font = SUB
        r += 1

    ws["C6"].number_format = "0"
    for rng, formula in [("C7", '"M,F"'), ("C8", '"N,Y"'), ("C9", '"S,R"')]:
        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws[rng])
    dv_age = DataValidation(type="whole", operator="between",
                            formula1=0, formula2=100, allow_blank=False)
    ws.add_data_validation(dv_age)
    dv_age.add(ws["C6"])

    # Everything below flows from a cursor rather than hardcoded row numbers.
    # The previous layout pinned each block to a literal row, so adding a
    # single warning silently pushed the footer into the block beneath it.
    r += 2

    def heading(text, colour=None):
        nonlocal r
        c = ws.cell(r, 2, text)
        c.font = Font(name=ARIAL, size=11, bold=True, color=colour or "000000")
        r += 1

    def line(text, font=None, fill=None, height=None, span=4):
        nonlocal r
        c = ws.cell(r, 2, text)
        c.font = font or BLACK
        if fill:
            c.fill = fill
        if height:
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = height
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=span)
        r += 1

    # This file is replaced every refresh, so freshness is the first thing he
    # needs to know and the easiest thing to forget. It goes above the usage
    # steps, not in a footer.
    heading("BEFORE YOU QUOTE ANYONE", "C00000")
    line(f"This copy was built on {datetime.now().strftime('%d %B %Y')}. "
         f"The newest premium in it takes effect {prem_date_max}.",
         font=Font(name=ARIAL, size=10, bold=True), fill=WARN_FILL, height=16)
    line("Download a fresh copy before each client meeting. The file is rebuilt "
         "whenever the Health Bureau publishes new data, and this one does not "
         "update itself.",
         font=Font(name=ARIAL, size=9), fill=WARN_FILL, height=16)
    line("Always the newest copy:  "
         "github.com/RianMehta21/insurance/releases/latest",
         font=Font(name=ARIAL, size=9, bold=True), fill=WARN_FILL, height=16)
    line("Do not save your own notes into this file. It is a tool, not a "
         "document: the next download replaces it entirely. Copy your shortlist "
         "into your own workbook instead.",
         font=Font(name=ARIAL, size=9), fill=WARN_FILL, height=16)
    r += 1

    heading("HOW TO USE")
    for s in [
        "1. Check the build date above. If it is more than a week old, download a fresh copy.",
        "2. Set the four yellow cells at the top.",
        "3. Go to the Compare sheet. Every premium column has already recalculated.",
        "4. Filter row 4: set Available = Open first, then narrow by ward, deductible, geography.",
        "5. Sort by 10-Year Avg, not First Year. See the warnings below.",
        "6. Copy the shortlisted rows into your own file for the client.",
    ]:
        line(s)
    r += 1

    heading("COLOUR LEGEND")
    for lbl, fill, desc in [
            ("Blue on yellow", INPUT_FILL, "You type here. These are the only cells to change."),
            ("Orange fill", EDIT_FILL, "Benefits sheet. Rebuilt each refresh; corrections go in manual_overrides.csv."),
            ("Black text", None, "Formula. Do not overwrite or the sheet stops recalculating."),
            ("Orange row", WARN_FILL, "Compare sheet. This plan cannot be sold to a new client.")]:
        c = ws.cell(r, 2, lbl)
        c.font = BLUE if fill is INPUT_FILL else BLACK
        if fill:
            c.fill = fill
        c.border = BOX
        ws.cell(r, 4, desc).font = SUB
        r += 1
    r += 1

    heading("READ BEFORE QUOTING", "C00000")
    for w in [
        "Age 75 and above: the source data stops at each insurer's new-application age ceiling, "
        "so the 10-year window is incomplete for 546 of 579 plans. Check the Years Avail column.",
        "The Available column shows whether a plan can still be sold to a NEW client. "
        "69 plans are de-registered, withdrawn or renewal-only. They price normally, so filter "
        "Available = Open before shortlisting.",
        "Premiums exclude the Insurance Authority levy, underwriting loading, and any discounts. "
        "They are portfolio-basis standard premiums, not quotes.",
        "Paying monthly typically costs about 8% more per year than annual. Not reflected here.",
        "Age-nearest-birthday insurers are approximated from whole-year age. "
        "Those rows show APPROX and need date of birth to be exact.",
        "Never rank across currencies. Filter to HKD or USD first.",
        "Deductible is blank for some Flexi plans because the plan document does not state it; "
        "it is set per policy. Confirm with the insurer before quoting those.",
    ]:
        line(w, font=Font(name=ARIAL, size=9), fill=WARN_FILL, height=26)
    r += 1

    ws.cell(r, 2, "Data refreshed").font = BLACK
    ws.cell(r, 3, datetime.now().strftime("%Y-%m-%d")).font = BLACK
    ws.cell(r + 1, 2, "Latest premium date in data").font = BLACK
    ws.cell(r + 1, 3, prem_date_max).font = BLACK
    ws.cell(r + 2, 2, "Source").font = BLACK
    ws.cell(r + 2, 3, "data.gov.hk / Health Bureau VHIS open data").font = SUB
    return ws


COMPARE_COLS = OrderedDict([
    ("Insurer", 26), ("Plan Name", 34), ("Type", 8), ("Available", 17),
    ("Plan Level", 30),
    ("Ward", 14), ("Deductible", 12), ("Geography", 18), ("Ccy", 6),
    ("First Year", 12), ("10-Year Total", 14), ("10-Year Avg", 13),
    ("Years Avail", 11), ("Age Basis", 11), ("Basis Used", 11),
    ("Smoker Rated", 12), ("Annual Limit", 13), ("Coinsurance %", 12),
    ("Cert No", 22), ("_tableage", 10), ("_key", 30), ("_row", 8),
])

# name -> 1-based column index, and name -> column letter.
#
# Everything below addresses columns through these maps rather than through
# literal numbers and letters. Inserting "Available" into the middle of the
# layout above previously meant hand-renumbering ws.cell(r, 13) and every $M /
# $R / $S reference inside the formula strings, where one missed substitution
# produces a workbook that opens cleanly and quietly reads the wrong column.
CIDX = {name: i for i, name in enumerate(COMPARE_COLS, start=1)}
CL = {name: get_column_letter(i) for name, i in CIDX.items()}


def sheet_compare(wb, catalog, n_prem_rows):
    ws = wb.create_sheet("Compare")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    ws["A1"] = "Compare"
    ws["A1"].font = TITLE
    ws["A2"] = ("Live. Driven by the yellow cells on the Client sheet. "
                "Filter with the arrows on row 4. Sort by 10-Year Avg.")
    ws["A2"].font = SUB

    for i, (name, width) in enumerate(COMPARE_COLS.items(), start=1):
        c = ws.cell(4, i, name)
        c.font = HDR
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width

    prem_last_row = n_prem_rows + 1
    keys = f"prem!$A$2:$A${prem_last_row}"
    grid = (f"prem!${get_column_letter(PREM_FIRST_COL)}$2:"
            f"${get_column_letter(PREM_LAST_COL)}${prem_last_row}")

    age = "Client!$C$6"
    gender = "Client!$C$7"
    smoker = "Client!$C$8"
    basis = "Client!$C$9"

    # The Benefits sheet carries a note in row 1 and its header in row 2, so
    # its data starts at row 3. Ranges anchored at row 2 lined up with each
    # other and so returned correct values, but silently excluded the last
    # plan in the catalog, which never received any scraped benefit data.
    ben_first = 3
    ben_last = len(catalog) + 2
    ben_keys = f"Benefits!$A${ben_first}:$A${ben_last}"

    # Column letters, resolved by name.
    L_cert = CL["Cert No"]
    L_age = CL["Age Basis"]
    L_basis = CL["Basis Used"]
    L_tage = CL["_tableage"]
    L_key = CL["_key"]
    L_rowno = CL["_row"]
    L_total = CL["10-Year Total"]
    L_years = CL["Years Avail"]

    def ben_lookup(ben_col):
        return (f'INDEX(Benefits!${ben_col}${ben_first}:${ben_col}${ben_last},'
                f'MATCH(${L_cert}{{r}},{ben_keys},0))')

    r = 5
    for row in catalog:
        cert = row["certification_no"]
        ws.cell(r, CIDX["Insurer"], row["insurer"]).font = BLACK
        ws.cell(r, CIDX["Plan Name"], row["plan_name"]).font = BLACK
        ws.cell(r, CIDX["Type"], row["plan_type"]).font = BLACK
        ws.cell(r, CIDX["Plan Level"], row["plan_level_raw"]).font = BLACK

        # Withdrawn and renewal-only plans price exactly like live ones, so
        # without this column they sort into a shortlist indistinguishably.
        av = ws.cell(r, CIDX["Available"], row.get("availability", ""))
        av.font = BLACK
        if row.get("sellable_new") != "Y":
            av.fill = WARN_FILL

        # Ward / deductible / geography: prefer the Benefits sheet (scraped or
        # hand-entered), fall back to what we parsed from the plan-level label.
        #
        # numeric=True emits the fallback bare rather than quoted. Quoting it
        # made Excel return the text "16000", which looks identical in the cell
        # but drops out of numeric sorts and "greater than" filters - on the
        # column most likely to be filtered.
        for name, ben_col, fallback, numeric in (
                ("Ward", "B", row["ward_type"], False),
                ("Deductible", "C", row["deductible_amount"], True),
                ("Geography", "D", row["geographical_coverage"], False)):
            if fallback == "" or fallback is None:
                fb = '""'
            elif numeric:
                fb = str(fallback)
            else:
                fb = '"{}"'.format(str(fallback).replace('"', '""'))
            idx = ben_lookup(ben_col).format(r=r)
            c = ws.cell(r, CIDX[name],
                        f'=IFERROR(IF({idx}="",{fb},{idx}),{fb})')
            c.font = BLACK
        ws.cell(r, CIDX["Deductible"]).number_format = money()

        ws.cell(r, CIDX["Ccy"], row["currency"]).font = BLACK
        ws.cell(r, CIDX["Age Basis"], row["age_counting_method"]).font = BLACK
        ws.cell(r, CIDX["Smoker Rated"], row["smoker_rated"]).font = BLACK
        ws.cell(r, CIDX["Cert No"], cert).font = Font(name=ARIAL, size=9,
                                                      color="808080")

        # helper: table age per insurer's own age-counting convention.
        # Clamped to the last age column: an age-next-birthday insurer at
        # client age 100 would otherwise index age 101 and return #REF!.
        ws.cell(r, CIDX["_tableage"],
                f'=MIN(IF(${L_age}{r}="N",{age}+1,{age}),{MAX_AGE})').font = BLACK

        # helper: basis actually used, with automatic fallback
        ws.cell(r, CIDX["Basis Used"],
                f'=IF(COUNTIF({keys},${L_cert}{r}&"|"&{basis}&"|"&{gender}&"|"'
                f'&{smoker})>0,{basis},IF({basis}="S","R","S"))').font = BLACK

        ws.cell(r, CIDX["_key"],
                f'=${L_cert}{r}&"|"&${L_basis}{r}&"|"&{gender}&"|"&{smoker}'
                ).font = BLACK
        ws.cell(r, CIDX["_row"],
                f'=IFERROR(MATCH(${L_key}{r},{keys},0),"")').font = BLACK

        col_first = f"${L_tage}{r}+1"
        col_last = f"MIN(${L_tage}{r}+10,{N_AGE_COLS})"
        rowno = f"${L_rowno}{r}"

        # First year. INDEX returns 0 on a blank cell, so gate on COUNT.
        ws.cell(r, CIDX["First Year"],
                f'=IF({rowno}="","",IF(COUNT(INDEX({grid},{rowno},{col_first}))=0,'
                f'"",INDEX({grid},{rowno},{col_first})))').number_format = money()
        ws.cell(r, CIDX["10-Year Total"],
                f'=IF({rowno}="","",IF(${L_years}{r}=0,"",'
                f'SUM(INDEX({grid},{rowno},{col_first}):'
                f'INDEX({grid},{rowno},{col_last}))))').number_format = money()
        ws.cell(r, CIDX["10-Year Avg"],
                f'=IF(${L_total}{r}="","",IF(${L_years}{r}=0,"",'
                f'${L_total}{r}/${L_years}{r}))').number_format = money()
        ws.cell(r, CIDX["Years Avail"],
                f'=IF({rowno}="",0,COUNT(INDEX({grid},{rowno},{col_first}):'
                f'INDEX({grid},{rowno},{col_last})))')

        for name, ben_col in (("Annual Limit", "E"), ("Coinsurance %", "F")):
            idx = ben_lookup(ben_col).format(r=r)
            ws.cell(r, CIDX[name], f'=IFERROR(IF(COUNT({idx})=0,"",{idx}),"")')
        ws.cell(r, CIDX["Annual Limit"]).number_format = money()

        for name in ("First Year", "10-Year Total", "10-Year Avg", "Years Avail"):
            ws.cell(r, CIDX[name]).font = BLACK
        r += 1

    last = r - 1
    ws.auto_filter.ref = f"A4:{get_column_letter(len(COMPARE_COLS))}{last}"

    # Hide the helper columns. They must stay inside the filter range for the
    # formulas to resolve, but nobody needs to look at them.
    for name in COMPARE_COLS:
        if name.startswith("_"):
            ws.column_dimensions[CL[name]].hidden = True
    return ws, last


def load_benefits(benefits_csv, overrides_csv):
    """Merge scraper output with hand-entered overrides. Overrides win.

    These live in CSV files, NOT in the workbook, because the workbook is
    regenerated on every refresh. Anything typed into the spreadsheet itself
    would be silently destroyed the next time the pipeline runs.
    """
    import csv as _csv
    merged = {}
    for path, tag in ((benefits_csv, "scraped"), (overrides_csv, "override")):
        if not path or not os.path.exists(path):
            continue
        with open(path, encoding="utf-8-sig") as f:
            for row in _csv.DictReader(f):
                cert = (row.get("certification_no") or "").strip()
                if not cert:
                    continue
                cur = merged.setdefault(cert, {})
                supplied = False
                # Accept the scraper's long column names AND plain short ones,
                # so a hand-written overrides file does not silently no-op.
                for src, dst in (("ward_restriction", "ward"),
                                 ("ward_type", "ward"),
                                 ("ward", "ward"),
                                 ("deductible_amount", "deductible"),
                                 ("deductible", "deductible"),
                                 ("geographical_coverage", "geography"),
                                 ("geography", "geography"),
                                 ("annual_benefit_limit", "annual_limit"),
                                 ("annual_limit", "annual_limit"),
                                 ("coinsurance_pct", "coinsurance"),
                                 ("coinsurance", "coinsurance"),
                                 ("lifetime_benefit_limit", "lifetime"),
                                 ("lifetime", "lifetime")):
                    v = (row.get(src) or "").strip()
                    if v:
                        cur[dst] = v
                        supplied = True

                # Credit the row only if it actually supplied something.
                #
                # manual_overrides.csv doubles as the to-do list, so it holds a
                # mostly-blank placeholder row for every plan still awaiting
                # hand entry. Keying off "does this plan have values by now"
                # instead of "did THIS row provide them" stamped 229 plans as
                # override when the numbers came from the scraper - and the
                # Source column is exactly what you would check to decide
                # whether a figure had been verified by a person.
                if supplied:
                    cur["tier"] = ("override" if tag == "override"
                                   else row.get("overall_tier", "scraped"))
    return merged


def sheet_benefits(wb, catalog, benefits=None):
    ws = wb.create_sheet("Benefits")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    ws["A1"] = ("Benefit fields from scrape_plans.py plus manual overrides. "
                "DO NOT EDIT HERE: this sheet is rebuilt from data/plan_benefits.csv "
                "and data/manual_overrides.csv every refresh, so edits made in Excel "
                "are lost. Put corrections in manual_overrides.csv instead.")
    ws["A1"].font = SUB
    ws.merge_cells("A1:H1")

    cols = [("Cert No", 22), ("Ward", 16), ("Deductible", 13), ("Geography", 20),
            ("Annual Limit", 14), ("Coinsurance %", 13), ("Lifetime Limit", 15),
            ("Source / Tier", 16)]
    for i, (name, w) in enumerate(cols, start=1):
        c = ws.cell(2, i, name)
        c.font = HDR
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = w

    benefits = benefits or {}
    for i, row in enumerate(catalog, start=3):
        cert = row["certification_no"]
        ws.cell(i, 1, cert).font = Font(name=ARIAL, size=9)
        b = benefits.get(cert, {})
        vals = [b.get("ward", ""), b.get("deductible", ""), b.get("geography", ""),
                b.get("annual_limit", ""), b.get("coinsurance", ""),
                b.get("lifetime", ""), b.get("tier", "")]
        for j, col in enumerate(range(2, 9)):
            c = ws.cell(i, col)
            v = vals[j]
            if v not in ("", None):
                try:
                    c.value = int(str(v).replace(",", ""))
                except ValueError:
                    c.value = v
            c.fill = EDIT_FILL
            c.font = BLACK
        ws.cell(i, 3).number_format = money()
        ws.cell(i, 5).number_format = money()
    return ws


def sheet_prem(wb, prem_rows):
    ws = wb.create_sheet("prem")
    ws.cell(1, 1, "key").font = HDR
    for a in range(N_AGE_COLS):
        ws.cell(1, PREM_FIRST_COL + a, a).font = HDR
    for i, row in enumerate(prem_rows, start=2):
        ws.cell(i, 1, row[0])
        for j, v in enumerate(row[1:]):
            if v is not None:
                ws.cell(i, PREM_FIRST_COL + j, v)
    ws.sheet_state = "hidden"
    return ws


def sheet_notes(wb, stats):
    ws = wb.create_sheet("Notes")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 78
    ws["A1"] = "Provenance and assumptions"
    ws["A1"].font = TITLE
    rows = [
        ("Source", "data.gov.hk, Health Bureau VHIS open data (4 JSON resources)"),
        ("Built", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Plan variants", stats["variants"]),
        ("Insurers", stats["insurers"]),
        ("Premium series rows", stats["prem_rows"]),
        ("Latest premium date", stats["prem_date_max"]),
        ("", ""),
        ("Age basis A", "Age last birthday (attained). Table age = client age."),
        ("Age basis N", "Age next birthday. Table age = client age + 1."),
        ("Age basis R", "Age nearest birthday. APPROXIMATED as client age; "
                        "exact resolution needs date of birth."),
        ("", ""),
        ("Basis S", "STANDALONE policy (獨立保單). Confirmed by the Health "
                    "Bureau data dictionary for the Standard Premium dataset. "
                    "Present for 530 of 579 variants."),
        ("Basis R", "RIDER / supplementary benefit attached to another policy "
                    "(附屬保單). Present for 184 variants, all from insurers that "
                    "also sell VHIS on top of a life policy. Of the 135 variants "
                    "priced on both bases: R is lower at every age on 87, "
                    "identical on 42, and HIGHER at one or more ages on 6. Quote "
                    "R only when the plan is genuinely being sold as a rider, "
                    "and do not assume it is the cheaper option."),
        ("Basis S/R caveat", "This standalone/rider reading is inferred from the "
                             "data, not read off a published data dictionary: "
                             "an 'S' series appears on 497 Flexi variants, which "
                             "rules out S meaning 'Standard Plan'. Confirm with "
                             "the Health Bureau before relying on it in a quote."),
        ("Basis fallback", "If the chosen basis has no table for a plan, the "
                           "other basis is used. See the Basis Used column."),
        ("", ""),
        ("Excluded from premiums", "Insurance Authority levy; underwriting "
                                   "premium loading; payment-mode loading "
                                   "(monthly is roughly +8%/yr); any discounts."),
        ("Age ceiling", "Source data stops at each insurer's new-application "
                        "age limit. Renewal-only ages are absent, so 10-year "
                        "figures truncate from client age 72."),
        ("Benefit fields", "Ward, deductible and geography are parsed from the "
                           "plan label where possible and otherwise must come "
                           "from the plan PDF. The Benefits sheet overrides."),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(i, 1, k).font = Font(name=ARIAL, size=10, bold=bool(k))
        c = ws.cell(i, 2, v)
        c.font = BLACK
        c.alignment = Alignment(wrap_text=True, vertical="top")
    return ws


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data-dir", default="current",
                    help="directory holding the four source JSON files")
    ap.add_argument("--out", default="VHIS_Compare.xlsx")
    ap.add_argument("--benefits", default="data/plan_benefits.csv")
    ap.add_argument("--overrides", default="data/manual_overrides.csv")
    args = ap.parse_args()

    std, flexi, prem, providers = bv.load(args.data_dir)
    prem_by_cert = {r["certification-no"]: r for r in prem}
    catalog, _ = bv.build_catalog(std, flexi, prem_by_cert, providers)
    catalog.sort(key=lambda r: (r["insurer"], r["plan_name"], r["plan_level_raw"]))

    prem_rows = build_premium_rows(prem_by_cert)
    prem_date_max = max(r["prem-date"] for r in prem)
    prem_date_max = (f"{prem_date_max[:4]}-{prem_date_max[4:6]}-{prem_date_max[6:]}")

    wb = Workbook()
    wb.remove(wb.active)

    sheet_client(wb, len(catalog), prem_date_max)
    sheet_compare(wb, catalog, len(prem_rows))
    benefits = load_benefits(args.benefits, args.overrides)
    sheet_benefits(wb, catalog, benefits)
    sheet_prem(wb, prem_rows)
    sheet_notes(wb, {
        "variants": len(catalog),
        "insurers": len({r["insurer"] for r in catalog}),
        "prem_rows": len(prem_rows),
        "prem_date_max": prem_date_max,
    })

    wb.save(args.out)
    print(f"wrote {args.out}")
    print(f"  compare rows : {len(catalog)}")
    print(f"  premium rows : {len(prem_rows)} (vs 234,464 in the long CSV)")
    print(f"  benefit rows : {len(benefits)} merged from scrape + overrides")


if __name__ == "__main__":
    main()
